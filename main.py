from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from typing import List, Dict, Any, Optional
import io
import re
import os
import xlrd  # 👈 pour les fichiers .xls

app = FastAPI(title="Excel BOQ Parser v2")

# CORS (à restreindre en prod)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mots-clés pour détecter les rôles de colonnes
HEADER_KEYWORDS: Dict[str, List[str]] = {
    "quantity": ["quantité", "qté", "qte", "qty", "hoeveelheid"],
    "unit": ["unité", "unit", "eenheid", "u", "u.", "un", "un."],
    "unit_price": ["pu", "prix unitaire", "unit price", "eenheidsprijs"],
    "amount": ["montant", "total", "totaal", "amount", "sommes"],
    "weight_kg": ["kg", "poids", "gewicht"],
}

ARTICLE_CODE_REGEX = re.compile(r"^\d{3}(\.\d+)*(\.[A-Z])?\.$")


# ---------- Helpers bas niveau ----------

def extract_visible_rows_from_active_sheet(
    content: bytes,
    file_extension: str,
    max_rows: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Ouvre un fichier Excel, prend la dernière feuille active (xlsx) ou
    la première feuille visible (xls), renvoie les lignes visibles sous
    forme JSON. Ignore les lignes cachées et les lignes totalement vides.
    """
    file_extension = file_extension.lower()

    # ----- Cas XLSX / XLSM / XLTX / XLTM : openpyxl -----
    if file_extension in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Impossible de lire le fichier Excel (openpyxl): {e}")

        ws = wb.active  # feuille active (dernière ouverte)
        sheet_name = ws.title

        rows_out: List[Dict[str, Any]] = []

        for row in ws.iter_rows():
            first_cell = row[0]
            row_idx = first_cell.row

            # Ligne cachée ?
            row_dim = ws.row_dimensions.get(row_idx)
            if row_dim is not None and row_dim.hidden:
                continue

            cells = []
            all_empty = True

            for cell in row:
                col_index = cell.column
                col_letter = get_column_letter(col_index)
                address = f"{col_letter}{cell.row}"

                if isinstance(cell, MergedCell):
                    value = None
                else:
                    value = cell.value

                if value is not None and str(value).strip() != "":
                    all_empty = False

                cells.append(
                    {
                        "address": address,
                        "col_index": col_index,
                        "col_letter": col_letter,
                        "row_index": cell.row,
                        "value": value,
                    }
                )

            if all_empty:
                continue

            rows_out.append(
                {
                    "row_index": row_idx,
                    "cells": cells,
                }
            )

            if max_rows is not None and len(rows_out) >= max_rows:
                break

        return {
            "sheet_name": sheet_name,
            "row_count": len(rows_out),
            "rows": rows_out,
        }

    # ----- Cas XLS (ancien format binaire) : xlrd -----
    elif file_extension == ".xls":
        try:
            book = xlrd.open_workbook(file_contents=content, formatting_info=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Impossible de lire le fichier XLS (xlrd): {e}")

        # choisir la première feuille visible
        sheet = None
        for sh in book.sheets():
            # visibility: 0 = visible, 1 = hidden, 2 = very hidden
            if getattr(sh, "visibility", 0) == 0:
                sheet = sh
                break
        if sheet is None:
            sheet = book.sheet_by_index(0)

        sheet_name = sheet.name
        rows_out: List[Dict[str, Any]] = []

        for r in range(sheet.nrows):
            # rowinfo_map: infos de format, dont hidden
            row_info = sheet.rowinfo_map.get(r)
            if row_info is not None and getattr(row_info, "hidden", 0):
                continue  # ligne cachée

            cells = []
            all_empty = True
            excel_row_index = r + 1  # 1-based pour rester cohérent avec openpyxl

            for c in range(sheet.ncols):
                cell_obj = sheet.cell(r, c)
                value = cell_obj.value

                if value not in (None, "") and str(value).strip() != "":
                    all_empty = False

                col_index = c + 1  # 1-based
                col_letter = get_column_letter(col_index)
                address = f"{col_letter}{excel_row_index}"

                cells.append(
                    {
                        "address": address,
                        "col_index": col_index,
                        "col_letter": col_letter,
                        "row_index": excel_row_index,
                        "value": value,
                    }
                )

            if all_empty:
                continue

            rows_out.append(
                {
                    "row_index": excel_row_index,
                    "cells": cells,
                }
            )

            if max_rows is not None and len(rows_out) >= max_rows:
                break

        return {
            "sheet_name": sheet_name,
            "row_count": len(rows_out),
            "rows": rows_out,
        }

    else:
        raise HTTPException(status_code=400, detail=f"Extension de fichier non supportée: {file_extension}")


def detect_column_roles(
    rows_preview: List[Dict[str, Any]],
    quantity_header_hint: Optional[str] = None,
    unit_header_hint: Optional[str] = None,
) -> Dict[str, str]:
    """
    Détecte les rôles de colonnes (quantity, unit, amount, weight_kg, …)
    à partir d'un aperçu de quelques lignes.

    - Détection auto via HEADER_KEYWORDS
    - Puis override via quantity_header_hint / unit_header_hint si fournis.
    """
    roles: Dict[str, set[str]] = {}

    # 1️⃣ Détection automatique via keywords
    for row in rows_preview:
        for cell in row["cells"]:
            raw = cell["value"]
            if raw is None:
                continue

            v = str(raw).strip().lower()
            if not v:
                continue

            for role, keywords in HEADER_KEYWORDS.items():
                if any(k in v for k in keywords):
                    col = cell["col_letter"]
                    roles.setdefault(role, set()).add(col)

    # On simplifie : une seule colonne par rôle
    simplified: Dict[str, str] = {role: sorted(cols)[0] for role, cols in roles.items()}

    # Helper interne : chercher une colonne via un hint texte
    def find_col_by_hint(hint: str) -> Optional[str]:
        hint_norm = hint.strip().lower()
        for row in rows_preview:
            for cell in row["cells"]:
                raw = cell["value"]
                if raw is None:
                    continue
                v = str(raw)
                v_norm = " ".join(v.replace("|", " ").split()).lower()
                if hint_norm in v_norm:
                    return cell["col_letter"]
        return None

    # 2️⃣ HINT QUANTITY (override)
    if quantity_header_hint:
        col = find_col_by_hint(quantity_header_hint)
        if col:
            simplified["quantity"] = col

    # 3️⃣ HINT UNIT (override)
    if unit_header_hint:
        col = find_col_by_hint(unit_header_hint)
        if col:
            simplified["unit"] = col

    return simplified


def detect_article_code(cells: List[Dict[str, Any]]) -> Optional[str]:
    for cell in cells:
        raw = cell["value"]
        if raw is None:
            continue
        v = str(raw).strip()
        if ARTICLE_CODE_REGEX.match(v):
            return v
    return None


def normalize_unit(u: Optional[str]) -> Optional[str]:
    if u is None:
        return None
    v = str(u).strip().lower()
    v = v.replace("m²", "m2").replace("m³", "m3")
    return v or None


def cell_by_col_letter(cells: List[Dict[str, Any]], col_letter: str) -> Optional[Dict[str, Any]]:
    for c in cells:
        if c["col_letter"] == col_letter:
            return c
    return None


def to_boq_line(row: Dict[str, Any], column_roles: Dict[str, str]) -> Dict[str, Any]:
    cells = row["cells"]

    article_code = detect_article_code(cells)

    # Unité
    unit = None
    unit_col = column_roles.get("unit")
    if unit_col:
        c = cell_by_col_letter(cells, unit_col)
        if c and c["value"] is not None:
            unit = normalize_unit(c["value"])

    # Quantité
    quantity: Optional[float] = None
    qty_col = column_roles.get("quantity")
    if qty_col:
        c = cell_by_col_letter(cells, qty_col)
        if c and isinstance(c["value"], (int, float)):
            quantity = float(c["value"])

    # Poids kg
    weight_kg: Optional[float] = None
    w_col = column_roles.get("weight_kg")
    if w_col:
        c = cell_by_col_letter(cells, w_col)
        if c and isinstance(c["value"], (int, float)):
            weight_kg = float(c["value"])

    # Heuristique simple pour savoir si c'est une vraie ligne de métré
    has_description = any(
        isinstance(c["value"], str) and len(str(c["value"]).strip()) > 5
        for c in cells
    )
    has_numeric = any(isinstance(c["value"], (int, float)) for c in cells)

    is_boq_line = bool(
        has_description
        and has_numeric
        and (article_code is not None or unit is not None or quantity is not None or weight_kg is not None)
    )

    return {
        "row_index": row["row_index"],
        "is_boq_line": is_boq_line,
        "article_code": article_code,
        "unit": unit,
        "quantity": quantity,
        "weight_kg": weight_kg,
    }


def summarize_boq_lines(boq_lines: List[Dict[str, Any]]) -> Dict[str, Any]:
    total_quantity = 0.0
    total_weight_kg = 0.0

    for line in boq_lines:
        if not line.get("is_boq_line"):
            continue

        q = line.get("quantity")
        if isinstance(q, (int, float)):
            total_quantity += float(q)

        w = line.get("weight_kg")
        if isinstance(w, (int, float)):
            total_weight_kg += float(w)

    return {
        "total_quantity": total_quantity,
        "total_weight_kg": total_weight_kg,
    }


# ---------- Endpoints FastAPI ----------

@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/parse_excel")
async def parse_excel(
    file: UploadFile = File(...),
    max_rows: Optional[int] = None,
):
    """
    Endpoint brut :
    - reçoit un fichier Excel (xlsx/xls)
    - lit la dernière feuille active ou première visible
    - retourne les lignes visibles en JSON
    """
    filename = file.filename or "uploaded"
    _, ext = os.path.splitext(filename)
    ext = ext.lower()

    if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"):
        raise HTTPException(
            status_code=400,
            detail="Le fichier doit être un Excel .xlsx / .xlsm / .xltx / .xltm / .xls",
        )

    content = await file.read()

    data = extract_visible_rows_from_active_sheet(content, ext, max_rows=max_rows)

    return {
        "filename": filename,
        "sheet_name": data["sheet_name"],
        "row_count": data["row_count"],
        "rows": data["rows"],
    }


@app.post("/parse_excel_transformed")
async def parse_excel_transformed(
    file: UploadFile = File(...),
    max_rows: Optional[int] = None,
    quantity_header_hint: Optional[str] = None,
    unit_header_hint: Optional[str] = None,
):
    """
    Endpoint transformé :
    - reçoit un fichier Excel (xlsx/xls)
    - lit la dernière feuille active / première visible
    - ignore les lignes cachées
    - détecte les rôles des colonnes (avec hints optionnels)
    - produit des lignes BOQ normalisées
    - renvoie aussi un résumé global
    """
    filename = file.filename or "uploaded"
    _, ext = os.path.splitext(filename)
    ext = ext.lower()

    if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"):
        raise HTTPException(
            status_code=400,
            detail="Le fichier doit être un Excel .xlsx / .xlsm / .xltx / .xltm / .xls",
        )

    content = await file.read()

    data = extract_visible_rows_from_active_sheet(content, ext, max_rows=max_rows)
    rows = data["rows"]

    rows_preview = rows[:30]
    column_roles = detect_column_roles(
        rows_preview,
        quantity_header_hint=quantity_header_hint,
        unit_header_hint=unit_header_hint,
    )

    boq_lines = [to_boq_line(row, column_roles) for row in rows]
    summary = summarize_boq_lines(boq_lines)

    return {
        "filename": filename,
        "sheet_name": data["sheet_name"],
        "row_count": data["row_count"],
        "column_roles": column_roles,
        "boq_lines": boq_lines,
        "summary": summary,
    }
